import os
import json
import requests
from openpyxl import load_workbook

EXCEL_INPUT_CELL = ""  # e.g. "A1", "C12", etc.
EXCEL_OUTPUT_CELL_1 = ""
EXCEL_OUTPUT_CELL_2 = ""

GOOGLE_MAPS_API_KEY = os.environ['GOOGLE_MAPS_API_KEY']
GOOGLE_MAPS_API_URL = 'https://maps.googleapis.com/maps/api/geocode/json'


def get_geocode(search_string):
    params = {
        'address': search_string,
        'key': GOOGLE_MAPS_API_KEY
    }
    req = requests.get(GOOGLE_MAPS_API_URL, params=params)
    res = req.json()
    return res.get("results", [])[0].get("geometry", {}).get("location", {})


def fetch_station(geocode):
    request_params = {
        "lat": geocode.get("lat"),
        "long": geocode.get("lng"),
        "number": "10",
        "ashrae_version": "2017"
    }
    url = "http://ashrae-meteo.info/request_places.php"

    resp = requests.post(url, data=request_params)
    resp_str = json.dumps(resp.json())  # handles encoding.
    resp_json = json.loads(resp_str)
    stations = resp_json.get("meteo_stations", [])
    selected_station = stations[0]
    return selected_station


def remove_bom(text):
    resp_json = None
    resp_str = json.dumps(text)
    resp_loaded = json.loads(resp_str.decode('utf-8-sig'))
    without_bom = resp_loaded[1:-3]
    resp_json = json.loads(without_bom)
    return resp_json


def fetch_weather_data(station_data):
    request_params = {
        "wmo": station_data.get("wmo"),
        "ashrae_version": "2017",
        "si_ip": "SI"
    }
    url = "http://ashrae-meteo.info/request_meteo_parametres.php"
    resp = requests.post(url, data=request_params)
    resp_json = remove_bom(resp.text)
    stations = resp_json.get('meteo_stations', [])
    station = stations[0]
    weather_data = {
        "cooling_DB_MCWB_2_DB": station.get('cooling_DB_MCWB_2_DB', 'n/a').encode('utf-8'),
        "n-year_return_period_values_of_extreme_DB_50_min": station.get('n-year_return_period_values_of_extreme_DB_50_min', 'n/a').encode('utf-8')
    }
    return weather_data


def read_excel_location(excel_filename):
    wb = load_workbook(excel_filename)
    ws = wb.active
    return ws[EXCEL_INPUT_CELL].value


def write_excel_weather(excel_filename, val_1, val_2):
    wb = load_workbook(excel_filename)
    ws = wb.active
    ws[EXCEL_OUTPUT_CELL_1] = val_1
    ws[EXCEL_OUTPUT_CELL_2] = val_2
    wb.save(excel_filename)
    return


def get_excel_filenames():
    excel_filenames = []
    wd = os.getcwd()
    files = os.listdir(wd)
    for f in files:
        if '.xlsx' in f:
            excel_filenames.append(f)
    return excel_filenames  


def excel_export(data):
    val_1 = data['cooling_DB_MCWB_2_DB']
    val_2 = data['n-year_return_period_values_of_extreme_DB_50_min']
    write_excel_weather(excel_filename, val_1, val_2)


def main():
    excel_filenames = get_excel_filenames()
    for f in excel_filenames:
        location = read_excel_location(f)
        geocode = get_geocode(location)
        station = fetch_station(geocode)
        weather_data = fetch_weather_data(station)
        excel_export(weather_data)


if __name__ == "__main__":

    main()
