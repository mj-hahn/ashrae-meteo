import os
import json
import requests
import mechanize
from openpyxl import load_workbook

EXCEL_INPUT_CELL = ""  # e.g. "A1", "C12", etc.
EXCEL_OUTPUT_CELL_1 = ""
EXCEL_OUTPUT_CELL_2 = ""

GOOGLE_MAPS_API_KEY = os.environ['GOOGLE_MAPS_API_KEY']
GOOGLE_MAPS_API_URL = 'https://maps.googleapis.com/maps/api/geocode/json'


def get_geocode(search_string):
    params = {
        'address': search_string,
        'key': GOOGLE_MAPS_API_KEY
    }
    req = requests.get(GOOGLE_MAPS_API_URL, params=params)
    res = req.json()
    return res.get("results", [])[0].get("geometry", {}).get("location", {})


def fetch_station(br, geocode):
    request_params = {
        "lat": geocode.get("lat"),
        "long": geocode.get("lng"),
        "number": "10",
        "ashrae_version": "2017"
    }
    url = "http://ashrae-meteo.info/request_places.php"
    new_request = mechanize.Request(
        url=url, data=request_params, method="POST")
    resp = br.open(new_request)
    resp_content = resp.read()
    resp_json = json.loads(resp_content)
    resp_str = json.dumps(resp_json, indent=4)

    stations = resp_json.get("meteo_stations", [])
    selected_station = stations[0]
    return selected_station


def fetch_weather_data(br, station_data):
    request_params = {
        "wmo": station_data.get("wmo"),
        "ashrae_version": "2017",
        "si_ip": "SI"
    }
    url = "http://ashrae-meteo.info/request_meteo_parametres.php"
    new_request = mechanize.Request(
        url=url, data=request_params, method="POST")
    resp = br.open(new_request)
    resp_content = resp.read()
    # resp contains ufeff, convert to unicode.
    resp = resp_content.decode('utf-8-sig')
    j_resp = json.loads(resp)

    stations = j_resp.get('meteo_stations', [])
    station = stations[0]

    weather_data = {
        "cooling_DB_MCWB_2_DB": station.get('cooling_DB_MCWB_2_DB', 'n/a'),
        "n-year_return_period_values_of_extreme_DB_50_min": station.get('n-year_return_period_values_of_extreme_DB_50_min', 'n/a')
    }
    return weather_data


def read_excel_location(excel_filename):
    wb = load_workbook(excel_filename)
    ws = wb.active  # or, specify the workboo name with: ws = wb['sheet name']
    return ws[EXCEL_INPUT_CELL].value


def write_excel_weather(excel_filename, val_1, val_2):
    wb = load_workbook(excel_filename)
    ws = wb.active  # or, specify the workboo name with: ws = wb['sheet name']
    ws[EXCEL_OUTPUT_CELL_1] = val_1
    ws[EXCEL_OUTPUT_CELL_2] = val_2
    wb.save(excel_filename)
    return


def get_excel_filename():
    excel_filename = None
    wd = os.getcwd()
    files = os.listdir(wd)
    for f in files:
        if '.xlsx' in f:
            excel_filename = f
    return excel_filename


def main(args):

    br = mechanize.Browser()  # init mechanize

    excel_filename = get_excel_filename()
    location = read_excel_location(excel_filename)
    geocode = get_geocode(location)
    station = fetch_station(br, geocode)
    weather_data = fetch_weather_data(br, station)

    val_1 = weather_data['cooling_DB_MCWB_2_DB']
    val_2 = weather_data['n-year_return_period_values_of_extreme_DB_50_min']
    write_excel_weather(excel_filename, val_1, val_2)


if __name__ == "__main__":

    main(args)
